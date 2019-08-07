from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
# from openpyxl import workbook  # 可以新建excel文件
from collections import namedtuple  # 导入namedtuple
from Class_16_Homework.homework_0520.config_class import do_config


class HandleExcel:
    """
    处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        # 打开excel文件（已存在）
        self.wb = load_workbook(self.filename)
        # 定位表单
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active
        self.cases_list = []  # 定义一个存放所有cases命名元组对象的空列表
        # 定位单元格
        # 获取表头信息
        sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        # 创建一个命名元组类
        self.Sheet = namedtuple('Sheet', sheet_head_tuple)

    def get_cases(self):
        """
        获取所有的测试用例
        :return:
        """
        # 每次遍历，返回某行所有单元格的值
        for case_data in self.ws.iter_rows(min_row=2, max_row=10, values_only=True):
            self.cases_list.append(self.Sheet._make(case_data))
            # self.cases_list.append(self.Sheet(*case_data))

        return self.cases_list

    def get_case(self, row):
        """
        获取某一条测试用例
        :param row: 行号
        :return:
        """
        if isinstance(row, int) and (1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            print('存入的行号有误，行号应为大于1的整数')

    def write_result(self, row, actual, result):
        """
        将测试用例实际结果和通过结果保存到excel
        :param row: 行号
        :param actual: 实际结果
        :param result: 通过结果
        :return:
        """
        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            # self.ws.cell(row=row, column=6, value=actual)
            self.ws.cell(row=row, column=do_config('excel', 'actual_col'), value=actual)
            # self.ws.cell(row=row, column=7, value=result)
            self.ws.cell(row=row, column=do_config('excel', 'result_col'), value=result)
            self.wb.save(self.filename)


# do_excel = HandleExcel(do_config('file path', 'cases_path'))
