import unittest

from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
# from openpyxl import workbook  # 可以新建excel文件
from collections import namedtuple  # 导入namedtuple

from ddt import ddt, data  # 导入ddt，ddt和data必须同时导入
# 导入数学计算的测试类
from Class_16_Homework.homework_0515.lemon_0419_07_homework import MathCalculate
# 导入配置文件类
from Class_16_Homework.homework_0515.lemon_0515_01_homework import HandleConfig


class HandleExcel:
    """
    处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        # 打开excel文件（已存在）
        self.wb = load_workbook(self.filename)
        # 定位表单
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active
        self.cases_list = []  # 定义一个存放所有cases命名元组对象的空列表
        # 定位单元格
        # 获取表头信息
        sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        # 创建一个命名元组类
        self.Sheet = namedtuple('Sheet', sheet_head_tuple)

    def get_cases(self):
        """
        获取所有的测试用例
        :return:
        """
        # 每次遍历，返回某行所有单元格的值
        for case_data in self.ws.iter_rows(min_row=2, max_row=10, values_only=True):
            self.cases_list.append(self.Sheet._make(case_data))
            # self.cases_list.append(self.Sheet(*case_data))

        return self.cases_list

    def get_case(self, row):
        """
        获取某一条测试用例
        :param row: 行号
        :return:
        """
        if isinstance(row, int) and (1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            print('存入的行号有误，行号应为大于1的整数')

    def write_result(self, row, actual, result):
        """
        将测试用例实际结果和通过结果保存到excel
        :param row: 行号
        :param actual: 实际结果
        :param result: 通过结果
        :return:
        """
        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            # self.ws.cell(row=row, column=6, value=actual)
            self.ws.cell(row=row, column=TestDivide.handle_config('excel', 'actual_col'), value=actual)
            # self.ws.cell(row=row, column=7, value=result)
            self.ws.cell(row=row, column=TestDivide.handle_config('excel', 'result_col'), value=result)
            self.wb.save(self.filename)


@ddt  # 在类的上一行加
class TestDivide(unittest.TestCase):
    """
    测试减法类
    """
    handle_config = HandleConfig()
    # handle_excel = HandleExcel('cases.xlsx')
    handle_excel = HandleExcel(handle_config('file path', 'cases_path'))
    cases_list = handle_excel.get_cases()

    @classmethod
    def setUpClass(cls):
        """
        重写父类的类方法，全部实例方法（用例）执行完只会被调用1次
        :return:
        """
        # cls.file_name = 'test_result.txt'
        cls.file_name = cls.handle_config('file path', 'log_path')
        print('打开【{}】文件'.format(cls.file_name))
        print("{:=^40s}".format("开始执行用例"))
        cls.file = open(cls.file_name, mode='a', encoding='utf8')
        cls.file.write("{:=^40s}\n".format("开始执行用例"))

    @classmethod
    def tearDownClass(cls):
        cls.file.write('{:=^40s}\n'.format('用例执行结束'))
        cls.file.close()
        # wb.save('cases.xlsx')

    @data(*cases_list)  # 遍历用例
    def test_case(self, case_value):
        """
        测试两个负数相除
        :return:
        """
        # data_namedtuple = cases_list.pop(0)
        case_id = case_value.case_id
        msg = case_value.title
        l_data = case_value.l_data
        r_data = case_value.r_data
        expect_result = case_value.expect_result
        actual_result = MathCalculate(l_data, r_data).divide()
        # 将实际结果写入excel
        # ws.cell(row=case_id + 1, column=6, value=actual_result)
        try:
            self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
        except AssertionError as e:
            print('具体异常为：{}'.format(e))
            # self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'Fail', e))
            self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, TestDivide.handle_config('msg', 'fail_result'), e))
            # ws.cell(row=case_id + 1, column=7, value='Fail')
            # self.handle_excel.write_result(case_id + 1, actual_result, 'Fail')
            self.handle_excel.write_result(case_id + 1, actual_result, TestDivide.handle_config('msg', 'fail_result'))
            raise e
        else:
            # self.file.write('{}，执行结果：{}\n'.format(msg, 'Pass'))
            self.file.write('{}，执行结果：{}\n'.format(msg, TestDivide.handle_config('msg', 'success_result')))
            # ws.cell(row=case_id + 1, column=7, value='Pass')
            # self.handle_excel.write_result(case_id + 1, actual_result, 'Pass')
            self.handle_excel.write_result(case_id + 1, actual_result, TestDivide.handle_config('msg', 'success_result'))


if __name__ == '__main__':
    unittest.main()
