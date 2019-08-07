import unittest

from Class_16_Homework.homework_0510.lemon_0419_07_homework import MathCalculate

from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
# from openpyxl import workbook  # 可以新建excel文件
from collections import namedtuple  # 导入namedtuple

from ddt import ddt, data  # 导入ddt，ddt和data必须同时导入

# 使用load_workbook来实现excel读写
# 打开excel文件（已存在）
wb = load_workbook('cases.xlsx')

# 定位表单
ws = wb.active

# 定位单元格
# 获取表头信息
# tuple(ws.iter_rows(max_row=1)) 和 ws['A1:G1'] 一样
sheet_head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]

# 封装用例，将获取的数据转化为字典
# 定义一个嵌套字典的列表来保存用例
cases_list = []
# 方法一：
# for data in ws.iter_rows(min_row=2, max_row=5, values_only=True):
#     cases_list.append({
#         'case_id':data[0],
#         'title': data[1],
#         'l_data': data[2],
#         'r_data': data[3],
#         'expected': data[4],
#         'actual': data[5],
#         'result': data[6],
#     })

# 方法二：
# for data in ws.iter_rows(min_row=2, max_row=5, values_only=True):
#     cases_list.append(dict(zip(sheet_head_tuple, data)))

# 方法三：namedtuple
# 定义一个命名元组类，首个参数为类名，第二个参数为类的属性名以空格或英文逗号隔开的字符串
# Sheet = namedtuple('Sheet', 'case_id title l_data r_data expected actual result')

# 第二个参数也可以用序列类型数据来传参
Sheet = namedtuple('Sheet', sheet_head_tuple)

# 创建一个对象（_make方法）
for case_data in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    cases_list.append(Sheet._make(case_data))


@ddt  # 在类的上一行加
class TestDivide(unittest.TestCase):
    """
    测试减法类
    """

    @classmethod
    def setUpClass(cls):
        """
        重写父类的类方法，全部实例方法（用例）执行完只会被调用1次
        :return:
        """
        cls.file_name = 'test_result.txt'
        print('打开【{}】文件'.format(cls.file_name))
        print("{:=^40s}".format("开始执行用例"))
        cls.file = open(cls.file_name, mode='a', encoding='utf8')
        cls.file.write("{:=^40s}\n".format("开始执行用例"))

    @classmethod
    def tearDownClass(cls):
        cls.file.write('{:=^40s}\n'.format('用例执行结束'))
        cls.file.close()
        wb.save('cases.xlsx')

    @data(*cases_list)  # 遍历用例
    def test_case(self, case_value):
    # def test_negatives(self):
        """
        测试两个负数相除
        :return:
        """
        # data_namedtuple = cases_list.pop(0)
        case_id = case_value.case_id
        msg = case_value.title
        l_data = case_value.l_data
        r_data = case_value.r_data
        expect_result = case_value.expect_result
        actual_result = MathCalculate(l_data, r_data).divide()
        # 将实际结果写入excel
        ws.cell(row=case_id + 1, column=6, value=actual_result)
        try:
            self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
        except AssertionError as e:
            print('具体异常为：{}'.format(e))
            self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
            ws.cell(row=case_id + 1, column=7, value='Fail')
            raise e
        else:
            self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
            ws.cell(row=case_id + 1, column=7, value='Pass')

    # def test_neg_pos(self):
    #     """
    #     测试前者负数、后者正数相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_neg_zero(self):
    #     """
    #     测试前者负数、后者为0相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_positives(self):
    #     """
    #     测试两个正数相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id+1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id+1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_pos_neg(self):
    #     """
    #     测试前者正数、后者负数相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_pos_zero(self):
    #     """
    #     测试前者为正数、后者为0相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_zero_neg(self):
    #     """
    #     测试前者为0、后者负数相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_zero_pos(self):
    #     """
    #     测试前者为0、后者正数相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')
    #
    # def test_zero_zero(self):
    #     """
    #     测试两个0相除
    #     :return:
    #     """
    #     data_namedtuple = cases_list.pop(0)
    #     case_id = data_namedtuple.case_id
    #     msg = data_namedtuple.title
    #     l_data = data_namedtuple.l_data
    #     r_data = data_namedtuple.r_data
    #     expect_result = data_namedtuple.expect_result
    #     actual_result = MathCalculate(l_data, r_data).divide()
    #     # 将实际结果写入excel
    #     ws.cell(row=case_id + 1, column=6, value=actual_result)
    #     try:
    #         self.assertEqual(actual_result, expect_result, msg='测试{}失败'.format(msg))
    #     except AssertionError as e:
    #         print('具体异常为：{}'.format(e))
    #         self.file.write('{}，执行结果：{}，具体异常为：{}\n'.format(msg, 'fail', e))
    #         ws.cell(row=case_id + 1, column=7, value='Fail')
    #         raise e
    #     else:
    #         self.file.write('{}，执行结果：{}\n'.format(msg, 'pass'))
    #         ws.cell(row=case_id + 1, column=7, value='Pass')


if __name__ == '__main__':
    unittest.main()