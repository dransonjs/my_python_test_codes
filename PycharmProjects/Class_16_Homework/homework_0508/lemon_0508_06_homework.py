# 使用openpyxl将所有用例数据读取出来
# 分别使用多种方法来封装测试用例数据（嵌套字典的列表和嵌套命名元组的列表）

from openpyxl import load_workbook
from collections import namedtuple

wb = load_workbook('cases.xlsx')

ws = wb.active

# 使用openpyxl将所有用例数据读取出来
for data in tuple(ws.iter_rows(max_row=9, max_col=7, values_only=True)):
    print(data)

# 分别使用多种方法来封装测试用例数据（嵌套字典的列表和嵌套命名元组的列表）
sheet_head = tuple(ws.iter_rows(max_row=1, values_only=True))[0]

# 方法一：使用嵌套字典的列表来封装用例
cases = []

for data in tuple(ws.iter_rows(min_row=2, max_row=9, max_col=7, values_only=True)):
    cases.append(dict(zip(sheet_head, data)))

# 方法二：使用嵌套命名元组的列表来封装用例
cases2 = []
Sheet = namedtuple('Sheet', sheet_head)

for data in tuple(ws.iter_rows(min_row=2, max_row=9, max_col=7, values_only=True)):
    cases2.append(Sheet._make(data))

pass