from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
from openpyxl import workbook  # 可以新建excel文件

# 使用load_workbook来实现excel读写
# 打开excel文件（已存在）
wb = load_workbook('cases.xlsx')
# 一个excel文件有哪些部分组成（文件名、表单、单元格）
# 第一个参数为文件名，返回一个Workbook对象（相当整个excel文件）

# 定位表单
# 方法一：
# ws = wb['multiple']  # ws为Worksheet对象（相当excel中的一个表单）

# 方法二：
ws = wb.active  # 获取第一个表单

# 定位单元格cell
# 方法一：处理指定的某个单元格
# cell1 = ws.cell(row=2, column=2)  #cell1为Cell对象（相当一个表单中的某个单元格）
# print(cell1.value)

# 方法二：每遍历一次，处理一个单元格
# Worksheet对象中有如下属性：
# max_row,min_row,max_column,min_column
# for row_index in range(ws.min_row+1, ws.max_row+1):
#     for col_index in range(ws.min_column, ws.max_column+1):
#         data = ws.cell(row=row_index, column=col_index).value
#         print('值为：{}\n类型为：{}\n'.format(data, type(data)))

# 方法三：Worksheet对象中的方法
# iter_rows：返回一个生成器，是由每一行的数据构成的元组组成的
# iter_columns：返回一个生成器，是由每一列的数据构成的元组组成的
# for row_tuple in ws.iter_rows(min_row=2):  # 每遍历一次，会将某一行的所有单元格对象组成的元组返回
#     print(row_tuple)
#     for cell1 in row_tuple:
#         data = cell1.value
#         print('值为：{}\n类型为：{}\n'.format(data, type(data)))

# 每遍历一次，会将某一行的所有单元格的值组成的元组返回
# for row_tuple in ws.iter_rows(min_row=2, values_only=True):
#     print(row_tuple)
#     for data in row_tuple:
#         print('值为：{}\n类型为：{}\n'.format(data, type(data)))

# 方法四：指定需要处理的所有单元格
# 返回的是嵌套元组的元组，内层元组是由每行的每个单元格的cell对象组成
sheet = ws['A2:G6']

for row in sheet:
    for cell in row:
        data = cell.value
        print('值为：{}\n类型为：{}\n'.format(data, type(data)))
pass

# 读数据不需要关闭，也不需要保存excel

# excel中的数字类型数据读取到python中也是数字类型，非数字类型（序列类型数据（列表、元组）、字典等）都被当成字符串
# 可以使用eval把excel的非数字类型转换成相应python数据类型
c6 = ws.cell(6, 3).value
c6_tuple = eval(c6)
print(c6_tuple, type(c6_tuple))

# 在excel单元格中保存python数据类型时，一定要满足python语法规范，且eval函数第一个参数必须是str，否则eval转换时会报错
# if isinstance(data1, (int, float, bool)):
#     print('无需转换')
# else:
#     try:
#         eval(data1)
#     except Exception as e:
#         print(e)
