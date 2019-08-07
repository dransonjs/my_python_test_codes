from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
from openpyxl import workbook  # 可以新建excel文件

# 使用load_workbook来实现excel读写
# 打开excel文件（已存在）
wb = load_workbook('cases.xlsx')

# 定位表单
ws = wb.active

# 定位单元格
# 获取表头信息
# tuple(ws.iter_rows(max_row=1)) 和 ws['A1:G1'] 一样
sheet_head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]

# 封装用例，将获取的数据转化为字典
# 定义一个嵌套字典的列表来保存用例
cases_list = []
# 方法一：
# for data in ws.iter_rows(min_row=2, max_row=5, values_only=True):
#     cases_list.append({
#         'case_id':data[0],
#         'title': data[1],
#         'l_data': data[2],
#         'r_data': data[3],
#         'expected': data[4],
#         'actual': data[5],
#         'result': data[6],
#     })

# 方法二：
for data in ws.iter_rows(min_row=2, max_row=5, values_only=True):
    cases_list.append(dict(zip(sheet_head_tuple, data)))

# 方法三：nametuple

pass
