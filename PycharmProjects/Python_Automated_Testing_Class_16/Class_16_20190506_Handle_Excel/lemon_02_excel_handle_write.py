from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
from openpyxl import workbook  # 可以新建excel文件

# 使用load_workbook来实现excel读写
# 打开excel文件（已存在）
wb = load_workbook('cases.xlsx')

# 定位表单
ws = wb.active  # 获取第一个表单

# 在单元格里写入数据
# 方法一：获取要写入的单元格
# 写入文件时，excel文件必须处于关闭状态
# ws.cell(2, 6).value = 8

# 方法二：
# ws.cell(2, 7, value='Pass')

# 方法三：添加一行内容
row_datas = ((6, '正数与负数相乘', 10, -3, 30, None, None),
             (7, '零与负数相乘', 0, -3, 0, None, None))

# 可以使用append方法来添加一行数据
for row in row_datas:
    ws.append(row)

# 保存文件(只要修改excel，都一定要保存)
wb.save('cases.xlsx')